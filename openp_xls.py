# # Работа с exel таблицами
# from openpyxl import Workbook
# wb = Workbook()
#
# # захватить активный лист
# ws = wb.active
#
# # Данные могут быть присвоены непосредственно ячейкам
# ws['A1'] = 42
#
# # Строки также могут быть добавлены
# ws.append([1, 2, 3])
#
# # Типы Python будут автоматически преобразованы
# import datetime
# ws['A2'] = datetime.datetime.now()
#
# # Сохранить файл
# wb.save("sample.xlsx")
import openpyxl

wb = openpyxl.reader.excel.load_workbook(filename='simple.xlsx', data_only=True)
print('Количество листов в книге:', wb.sheetnames)
wb.active = 1
print('Активный лист:', wb.active)
sheet = wb.active
# print(sheet['A1'].value)
for i in range(1, 11):
    print(sheet['A'+str(i)].value, sheet['B'+str(i)].value, sheet['C'+str(i)].value)
